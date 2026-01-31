#from singleinvoice.mailtest import *
import cx_Oracle
#from xlutils.copy import copy
#from copy import copy
#import xlrd
#import xlwt as xw
import datetime
#import xlsxwriter
import datetime
import time
import csv
import zipfile
import os.path
from os import path
from configparser import ConfigParser
import openpyxl as xl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import logging
import subprocess
import random
import string
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText




def xlParams():
    global lngInvoiceNo
    global xlApp
    global xlbkOutput
    global xlPosition
    global xlFileName
    global xlExternalCustFileName
    global xlCurrency
    global xlBillFormat
    global xlBillCurrency
    global sheet
    global Format
    global workbook
    global xlbkOutput_name
    global invoice_dt
    global Root_Directory
    global Template_directory
    global Prod_directory
    global Base_directory
    global Log_directory
    global Password_directory
    global zip_directory
    global Db_username
    global Db_Password
    global Db_Database
    global logging
    global log_level
    global file_handler
    global csv
    global pdf
    global encrypt
    global conn
    global cur
    global MailServer
    global MailFromAddress
    global MailSubject    
    global MailBody
    global encryptMailBody
    global PasswordMailSubject
    global PasswordMailBody
    global Prod_directory
    global Password_directory
    

    


def mthdGetCustInvoice (inv_type,strExternal_id,invoice_num,boolUsesCSV,boolUsesPDF,boolUsesEncrypt,cust_email,user_name):
    xlParams.xlFileName = ""
    xlParams.lngInvoiceNo = invoice_num    
    xlParams.xlCurrency = ""
    xlParams.invoice_dt=""
    cwd = os.getcwd() 
    #print(cwd)   
    xlParams.Template_directory = "BOD\\Template\\"
    xlParams.Prod_directory = "BOD\\Production\\"
    xlParams.Log_directory = "BOD\\Logs\\"
    xlParams.Password_directory = "BOD\\PasswordProtect\\"
    parser = ConfigParser ()    
    parser.read("templates\\config.ini")    
    xlParams.Db_username  = parser.get('DB', 'DB_USERNAME')
    xlParams.Db_Password  = parser.get('DB', 'DB_PASSWORD')
    xlParams.Db_Database  = parser.get('DB', 'DB_DATABASE')
    xlParams.Root_Directory = parser.get('DB', 'ROOT_DIR')
    xlParams.zip_directory = parser.get('DB', 'ZIP_DIR')
    xlParams.Prod_directory = "BOD\\Production\\"
    xlParams.Password_directory = "BOD\\PasswordProtect\\"    
    xlParams.MailServer  = parser.get('MAIL', 'MailServer')
    xlParams.MailFromAddress  = parser.get('MAIL', 'MailFromAddress')
    xlParams.MailSubject  = parser.get('MAIL', 'MailSubject')
    xlParams.MailBody = parser.get('MAIL', 'MailBody')
    xlParams.encryptMailBody = parser.get('MAIL', 'EncryptMailBody')    
    xlParams.PasswordMailSubject = parser.get('MAIL', 'PasswordMailSubject')
    xlParams.PasswordMailBody = parser.get('MAIL', 'PasswordMailBody')
    
    

    xlParams.logger = logging.getLogger(__name__)
    xlParams.log_level= parser.get('LOG', 'LOG_LEVEL')
    xlParams.logger.setLevel(xlParams.log_level)    
    os.chdir(xlParams.Root_Directory)
    xlParams.csv=""
    xlParams.pdf=""
    xlParams.encrypt=""
    if(boolUsesCSV ==1):
        xlParams.csv="True"
    if(boolUsesPDF ==1):
        xlParams.pdf="True"
    if(boolUsesEncrypt==1):
        xlParams.encrypt="True"

    


    
    
    
    mthdOutputToExcel_status = mthdOutputToExcel(cust_email,user_name,boolUsesEncrypt)
    xlParams.logger.info("Ths status of the method mthdOutputToExcel is " + str(mthdOutputToExcel_status))
    os.chdir(cwd)
    if (mthdOutputToExcel_status != 'Successful'):
        close_logging()
        return mthdOutputToExcel_status
    else:
        close_logging()
        return mthdOutputToExcel_status

def mthdOutputToExcel(cust_email,user_name,boolUsesEncrypt):
    mthdGetCustomerDetails_status = mthdGetCustomerDetails()
    xlParams.logger.info("Ths status of the method mthdGetCustomerDetails is " + str(mthdGetCustomerDetails_status))
    if (mthdGetCustomerDetails_status != 'Successful'):
        return mthdGetCustomerDetails_status
    mthdGetChargeDetails_status = mthdGetChargeDetails()
    xlParams.logger.info("Ths status of the method mthdGetChargeDetails is " + str(mthdGetChargeDetails_status))
    if (mthdGetChargeDetails_status != 'Successful'):
        return mthdGetChargeDetails_status
    mthdGetServiceSummary_status = mthdGetServiceSummary()
    xlParams.logger.info("Ths status of the method mthdGetServiceSummary is " + str(mthdGetServiceSummary_status))
    if (mthdGetServiceSummary_status != 'Successful'):
        return mthdGetServiceSummary_status
    mthdGetPaymentsAndAdj_status = mthdGetPaymentsAndAdj()
    xlParams.logger.info("Ths status of the method mthdGetPaymentsAndAdj is " + str(mthdGetPaymentsAndAdj_status))
    if (mthdGetPaymentsAndAdj_status != 'Successful'):
        return mthdGetPaymentsAndAdj_status
    mthdGetCCs_status = mthdGetCCs()
    xlParams.logger.info("Ths status of the method mthdGetCCs is " + str(mthdGetCCs_status))
    if (mthdGetCCs_status != 'Successful'):
        return mthdGetCCs_status
    mthdGetDiscounts_status = mthdGetDiscounts()
    xlParams.logger.info("Ths status of the method mthdGetDiscounts is " + str(mthdGetDiscounts_status))
    if (mthdGetDiscounts_status != 'Successful'):
        return mthdGetDiscounts_status
    mthdGetServiceSummaryTotals_status = mthdGetServiceSummaryTotals()
    xlParams.logger.info("Ths status of the method mthdGetServiceSummaryTotals is " + str(mthdGetServiceSummaryTotals_status))
    if (mthdGetServiceSummaryTotals_status != 'Successful'):
        return mthdGetServiceSummaryTotals_status
    mthdOutputDetails_status = mthdOutputDetails()
    xlParams.logger.info("Ths status of the method mthdOutputDetails is " + str(mthdOutputDetails_status))
    if (mthdOutputDetails_status != 'Successful'):
        return mthdOutputDetails_status
    if (xlParams.csv == "True"):
            writeinto_csv_status = writeinto_csv()
            xlParams.logger.info("Ths status of the method writeinto_csv is " + str(writeinto_csv_status))
            if (mthdOutputDetails_status != 'Successful'):
                return mthdOutputDetails_status
    if (xlParams.pdf == "True"):
        create_pdf_status = create_pdf(cust_email,user_name,boolUsesEncrypt)
        xlParams.logger.info("Ths status of the method create_pdf is " + str(create_pdf_status))
        if(create_pdf_status != "Successful"):
            return create_pdf_status
        
            

    mthdSaveExcelSheet_status = mthdSaveExcelSheet(xlParams.csv,cust_email)
    xlParams.logger.info("Ths status of the method mthdSaveExcelSheet is " + str(mthdSaveExcelSheet_status))
    if (mthdSaveExcelSheet_status != 'Successful'):
        return mthdSaveExcelSheet_status
    elif (mthdSaveExcelSheet_status == 'Successful' and xlParams.pdf == "True" ):
        return "pdfSuccessful"
    else:
        return mthdSaveExcelSheet_status
    
    
        


           
            
    
    
    

def mthdGetCustomerDetails():
    try:
        #print("Method mthdGetCustomerDetails started")
        xlParams.xlbkOutput_name = ""
        xlParams.conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        xlParams.cur = xlParams.conn.cursor()
        #n_invoice = 25991843
        rsCustomerDetail=[]
        PO_Display=""
        Mandate_Id =""
        Tax_Exempt = ""
        rsCustomerDetails=xlParams.conn.cursor()
        
        xlParams.cur.callproc("bod_create_types.BOD_GET_CUST_DETAILS",[xlParams.lngInvoiceNo,rsCustomerDetails])
        #print(rsCustomerDetails)
        for cus_row in rsCustomerDetails:
            rsCustomerDetail=cus_row
            #print (rsCustomerDetail)
        
            
        if(len(rsCustomerDetail)>0):
            xlParams.xlBillFormat = rsCustomerDetail[16]
            v_currency =""
            v_currency = rsCustomerDetail[15]
            if v_currency== 1:
                xlParams.xlBillCurrency = "IEP"
                
            elif v_currency== 2:
                xlParams.xlBillCurrency = "GBP"
            elif v_currency== 4:
                xlParams.xlBillCurrency = "USD"
            elif v_currency== 3:
                xlParams.xlBillCurrency = "EURO"
            else:
                xlParams.xlBillCurrency = "EURO"
            
            xlParams.xlFileName = rsCustomerDetail[9]  + "-" + str(xlParams.lngInvoiceNo) + "-" + str(rsCustomerDetail[8].strftime('%d%m%Y'))
            xlParams.xlbkOutput_name = "BOD_Template_" + xlParams.xlBillCurrency +"." +"xlsx"
            xlParams.file_handler = logging.FileHandler(xlParams.Root_Directory + xlParams.Log_directory + xlParams.xlFileName + '.log',mode='w')
            formatter    = logging.Formatter('%(asctime)s : %(levelname)s : %(name)s : %(message)s')
            xlParams.file_handler.setFormatter(formatter)
            xlParams.logger.addHandler(xlParams.file_handler)
            xlParams.logger.debug("The Excel Template Selected is " + xlParams.xlbkOutput_name)
            xlParams.invoice_dt = rsCustomerDetail[8]            
            xlExternalCustFileName = str(xlParams.lngInvoiceNo) + "_" + str(rsCustomerDetail[8].strftime('%d%m%Y'))
            xlParams.logger.debug("The path is " + xlParams.Root_Directory + xlParams.Template_directory + xlParams.xlbkOutput_name)
            file_name = xlParams.Root_Directory + xlParams.Template_directory + xlParams.xlbkOutput_name
            xlParams.workbook = xl.load_workbook(filename=file_name)
            #logger.debug("Excel workbook loaded successfully")
            xlParams.sheet = xlParams.workbook.worksheets[0]
            xlParams.logger.debug("Worksheet loaded successfully")            
            
            #png_loc = xlParams.Root_Directory + xlParams.Template_directory + 'bti_image.png'
            #xlParams.logger.debug(xlParams.Root_Directory + xlParams.Template_directory + 'bti_image.png')
            #my_png = xl.drawing.image.Image(png_loc)
            #xlParams.sheet.add_image(my_png, "A3")
            xlParams.xlPosition = 16
            
            if ( rsCustomerDetail[0] is not None):
                if(rsCustomerDetail[0].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)]= rsCustomerDetail[0].strip()
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            xlParams.sheet["E16"] =  "Invoice Number: "+ str(rsCustomerDetail[7])
            
            if ( rsCustomerDetail[1] is not None):
                if(rsCustomerDetail[1].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)] = rsCustomerDetail[1]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            xlParams.sheet["E17"]  = "Statement Date: "+ str(rsCustomerDetail[8].strftime('%d/%m/%Y'))
            
            if (rsCustomerDetail[2] is not None):
                if(rsCustomerDetail[2].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)] = rsCustomerDetail[2]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            xlParams.sheet["E18"] = "Customer Account No: "+ str(rsCustomerDetail[9])            
            xlParams.sheet["E61"] = str(rsCustomerDetail[9])
            if (rsCustomerDetail[3] is not None):
                if(rsCustomerDetail[3].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)]= rsCustomerDetail[3]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            if ( rsCustomerDetail[4] is not None):
                if(rsCustomerDetail[4].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)]= rsCustomerDetail[4]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            if (rsCustomerDetail[5] is not None):
                if(rsCustomerDetail[5].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)]= rsCustomerDetail[5]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            if (rsCustomerDetail[6] is not None):
                if(rsCustomerDetail[6].strip()):
                    xlParams.sheet["B" + str(xlParams.xlPosition)]= rsCustomerDetail[6]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition +1
            xlParams.sheet["E19"] = "Invoice Period: See Summary Details"
            xlParams.sheet["E19"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B42"] = "IMPORTANT Service Announcement: we are withdrawing service to our 1850, 1890 and 076 numbers on 31 October 2021"
            #xlParams.sheet["B42"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B43"] = "This is a Service Termination Notice for service to our 1850, 1890 and 076 numbers."
            #xlParams.sheet["B43"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B44"] = "As directed by ComReg, Ireland’s Communications Regulator, all 1850, 1890 and 076 numbers are to be withdrawn from service."
            #xlParams.sheet["B44"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B45"] = "BT will cease services to these numbers on 31 October 2021."
            #xlParams.sheet["B45"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B46"] = "This means that callers to your affected number(s) after that date will be answered by a network announcement advising the number is no longer in use."
            #xlParams.sheet["B46"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B47"] = "If you have been provided with new numbers (such as starting with 0818 or 1800) please start using these instead of 1850, 1890 or 076 as soon as possible."
            #xlParams.sheet["B47"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B48"] = "To find out more, see comreg.ie/ngn. To discuss alternatives, please call your BT account manager or email us at clientservices-ire@bt.com."
            #xlParams.sheet["B48"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B51"] = "For any queries relating to your bill please call us on our"
            #xlParams.sheet["B51"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B52"] = "Customer Care line 1800 924 925 or"
            #xlParams.sheet["B52"].font = Font(name='Century Gothic',size=10,bold=False)
            #xlParams.sheet["B53"] = "email billingsupport@btireland.ie"
            #xlParams.sheet["B53"].font = Font(name='Century Gothic',size=10,bold=False)




            
            if rsCustomerDetail[17] is not None:
                if ( len(rsCustomerDetail[17]) >0):
                    xlParams.sheet["E20"] = "Mandate Id: "+ str(rsCustomerDetail[17])
                    xlParams.sheet["E20"].font = Font(name='Century Gothic',size=10,bold=False)
                    Mandate_Id ="True"
            
                
            
              
            if rsCustomerDetail[18] is not None:
                if (len(rsCustomerDetail[18].strip()) > 0):
                    #print("Going inside")
                    if(len(rsCustomerDetail[18]) > 48):
                        if(Mandate_Id =="True"):
                            xlParams.sheet["E21"]= "PO Number: "+ str(rsCustomerDetail[18][0:48])
                            xlParams.sheet["E21"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
                        else:
                            xlParams.sheet["E20"]= "PO Number: "+ str(rsCustomerDetail[18][0:48])
                            xlParams.sheet["E20"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"


                    else :
                        if(Mandate_Id =="True"):
                            xlParams.sheet["E21"] = "PO Number: "+ str(rsCustomerDetail[18])
                            xlParams.sheet["E21"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
                        else:
                            xlParams.sheet["E20"] = "PO Number: "+ str(rsCustomerDetail[18])
                            xlParams.sheet["E20"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
            
            if rsCustomerDetail[19] is not None:            
                if ((rsCustomerDetail[18] is None or len(rsCustomerDetail[18].strip()) <= 0) and len(rsCustomerDetail[19].strip()) > 0 ):
                    if(Mandate_Id =="True"):
                        if(len(rsCustomerDetail[19]) > 48):
                            xlParams.sheet["E21"]= "PO Number: "+ str(rsCustomerDetail[19][0:48])
                            xlParams.sheet["E21"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
                        else:
                            xlParams.sheet["E20"]= "PO Number: "+ str(rsCustomerDetail[19][0:48])
                            xlParams.sheet["E20"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"


                    else :
                        if(Mandate_Id =="True"):
                            xlParams.sheet["E21"] = "PO Number: "+ str(rsCustomerDetail[19])
                            xlParams.sheet["E21"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
                        else:
                            xlParams.sheet["E20"]= "PO Number: "+ str(rsCustomerDetail[19][0:48])
                            xlParams.sheet["E20"].font = Font(name='Century Gothic',size=10,bold=False)
                            PO_Display = "True"
                        
            if rsCustomerDetail[12]  is not None:      
                if (len(rsCustomerDetail[12].strip()) >0):
                    if (PO_Display == "True" and Mandate_Id =="True" ):
                        xlParams.sheet["E22"] = "Tax Exemption Number: " + str(rsCustomerDetail[12])
                        xlParams.sheet["E22"].font = Font(name='Century Gothic',size=10,bold=False)
                        Tax_Exempt = "True"
                    else:
                        xlParams.sheet["E21"] = "Tax Exemption Number: " + str(rsCustomerDetail[12])
                        xlParams.sheet["E21"].font = Font(name='Century Gothic',size=10,bold=False)
                        Tax_Exempt = "True"
                        
            if rsCustomerDetail[9]  is not None:            
                if (rsCustomerDetail[9][0:3] =='CAR'):
                    #print("CAR found")
                    if(Tax_Exempt == "True" and PO_Display == "True" and Mandate_Id =="True"):
                        xlParams.sheet["E23"] = "Payment due 28 days from date of invoice unless otherwise agreed in writing"
                        xlParams.sheet["E23"].font = Font(name='Century Gothic',size=10,bold=False)
                    else:
                        xlParams.sheet["E22"] = "Payment due 28 days from date of invoice unless otherwise agreed in writing"
                        xlParams.sheet["E22"].font = Font(name='Century Gothic',size=10,bold=False)

                else:
                    if(Tax_Exempt == "True" and PO_Display == "True" and Mandate_Id =="True"):
                        xlParams.sheet["E23"] = "Payment due by: " + str(rsCustomerDetail[13].strftime('%d/%m/%Y'))
                        xlParams.sheet["E23"].font = Font(name='Century Gothic',size=10,bold=False)
                    else :
                        xlParams.sheet["E22"] = "Payment due by: " + str(rsCustomerDetail[13].strftime('%d/%m/%Y'))                   
                        xlParams.sheet["E22"].font = Font(name='Century Gothic',size=10,bold=False)
                    
                
            
            IsCustomerUsingEuroCurrency=["AT","BE","BG","CY","CZ","DK","EE","FI","FR","DE","EL","SE","HU","IT","LV","LT","LU","MT","NL",
                            "PL","PT","RO","SK","SI","ES"]
            if (rsCustomerDetail[12][0:2] in IsCustomerUsingEuroCurrency):
                xlParams.sheet["A55"] =  "                                        Not Subject to Irish VAT : reverse-Charge applies as per Art 196 Council Directive"
                xlParams.sheet["A55"].font = Font(name='Century Gothic',size=10,bold=False)
                xlParams.sheet["A56"] = "                                        2006/112/EC or Art 138 Council Directive 2006/112/EC refers."
                xlParams.sheet["A56"].font = Font(name='Century Gothic',size=10,bold=False)
                
            if (rsCustomerDetail[12][0:2] == "GB"):
                xlParams.sheet["A56"] = "                                        Non-EU - Not Subject to Irish VAT."
                xlParams.sheet["A56"].font = Font(name='Century Gothic',size=10,bold=False)
            
            
                    
            
            if (rsCustomerDetail[9][0:3] =="CAR" or
                rsCustomerDetail[9][0:2] =="PG" or
                rsCustomerDetail[9][0:2] =="IC" or
                rsCustomerDetail[9][0:2] =="99" or
                rsCustomerDetail[9][0:2] =="00" or
                rsCustomerDetail[9][0:2] =="01" or
                rsCustomerDetail[9][0:2] =="02" or
                rsCustomerDetail[9][0:2] =="03"):
                    #print("List found")
                    xlParams.sheet["E13"]= "VAT Registration No:                    IE6411141M"
                    #xlParams.sheet["E65"]= "Sort Code:                                 985010"
                    #xlParams.sheet["E66"] = "Account No:                              12366365"
                    #xlParams.sheet["B65"]= "Ulster Bank"
                    #xlParams.sheet["B66"]= "33 College Green"
                    #xlParams.sheet["B67"]= "Dublin 2"
                    #xlParams.sheet["B68"]= ""
            if (rsCustomerDetail[9][0:3] =="CP" or
                rsCustomerDetail[9][0:2] =="CP" or
                rsCustomerDetail[9][0:1] =="N"):
                   #print("List found")
                    xlParams.sheet["E13"] = "VAT Registration No:                    IE6411141M"
                    #xlParams.sheet["E65"] = "Sort Code:                                 985010"
                    #xlParams.sheet["E66"] = "Account No:                              12366365"
                    #xlParams.sheet["B65"] = "Ulster Bank"
                    #xlParams.sheet["B66"]= "33 College Green"
                    #xlParams.sheet["B67"]= "Dublin 2"
                    #xlParams.sheet["B68"]= ""
            if (rsCustomerDetail[9][0:2] =='IR' or 
                rsCustomerDetail[9][0:3] =='IR' or
                rsCustomerDetail[9][0:1] =='B'):
                    #print("List found")
                    xlParams.sheet["E13"]= "VAT Registration No:                    IE6411141M"
                    #xlParams.sheet["E65"]= "Sort Code:                                 985010"
                    #xlParams.sheet["E66"]= "Account No:                              12366365"
                    #xlParams.sheet["B65"]= "Ulster Bank"
                    #xlParams.sheet["B66"]= "33 College Green"
                    #xlParams.sheet["B67"]= "Dublin 2"
                    #xlParams.sheet["B68"]= ""
            if (xlParams.xlBillFormat ==9 ):
                #print("List found")
                xlParams.sheet["E13"]= "VAT Registration No:                      IE6411141M"
                #xlParams.sheet["E65"]= "Sort Code:                                 985010"
                #xlParams.sheet["E66"]= "Account No:                              12366365"
                #xlParams.sheet["B65"]= "Ulster Bank"
                #xlParams.sheet["B66"]= "33 College Green"
                #xlParams.sheet["B67"]= "Dublin 2"
                #xlParams.sheet["B68"]= ""
            if (xlParams.xlBillCurrency =="USD" ):
                #print("USD CURRENCY")
                xlParams.sheet["E13"]="VAT Registration No:                    IE6411141M"
                #xlParams.sheet["E65"]="Sort Code:                               98-50-05"
                #xlParams.sheet["E66"]="Account No:                              10121934"
                #xlParams.sheet["E67"]="BIC/SWIFT Code:                          ULSBIE2D"
                #xlParams.sheet["E68"]="IBAN:                      IE42ULSB98500510121934"
                #xlParams.sheet["E69"]="                                       BT Ireland"
                #xlParams.sheet["B65"]="Ulster Bank Ltd"
                #xlParams.sheet["B66"]="Georges Quay"
                #xlParams.sheet["B67"]="Dublin 2"
                #xlParams.sheet["B68"]=""
                
            if (xlParams.xlBillCurrency =="GBP" ):
                #print("USD CURRENCY")
                xlParams.sheet["E13"] = "VAT Registration No:            IE6411141M"
                #xlParams.sheet["E65"]= "Sort Code:                        98-50-05"
                #xlParams.sheet["E66"] = "Account No:                       10121850"
                #xlParams.sheet["E67"] = "BIC/SWIFT Code:                   ULSBIE2D"
                #xlParams.sheet["E68"] = "IBAN:               IE79ULSB98500510121850"
                #xlParams.sheet["E69"] = "                                BT Ireland"
                #xlParams.sheet["B65"] = "Ulster Bank Ltd"
                #xlParams.sheet["B66"] = "Georges Quay"
                #xlParams.sheet["B67"] = "Dublin 2"
                #xlParams.sheet["B68"] = ""
            
            if (rsCustomerDetail[14] != 1 ):
                xlParams.sheet["C51"] = ""
        else:
            xlParams.file_handler = logging.FileHandler(xlParams.Root_Directory + xlParams.Log_directory + "Error_" +str(xlParams.lngInvoiceNo) + '.log',mode='w')
            formatter    = logging.Formatter('%(asctime)s : %(levelname)s : %(name)s : %(message)s')
            xlParams.file_handler.setFormatter(formatter)
            xlParams.logger.addHandler(xlParams.file_handler)
            xlParams.logger.warning("No Data Found for Invoice in Procedure bod_create_types.BOD_GET_CUST_DETAILS")
            return "No Data Found for Invoice"     
       
        #cur.close()
        rsCustomerDetails.close()
        #conn.close()
        
        return "Successful"
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetCustomerDetails " + str(e))
        return e
    
def mthdGetChargeDetails():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #n_invoice = 25991843
        rsChargeDetail=[]
        rsChargeDetails=xlParams.conn.cursor()
        
        
        xlParams.cur.callproc("bod_create_types.BOD_GET_CHARGE_DETAILS",[xlParams.lngInvoiceNo,rsChargeDetails])
        for chg_row in rsChargeDetails:
            rsChargeDetail=chg_row
            
        #if(len(rsChargeDetail)>0):
            #xlParams.sheet["D34D34"] = rsChargeDetail[0]
        #if (rsChargeDetail[1] > 0 ):
            #xlParams.sheet["C35"] = "Payment Reversal"
        #xlParams.sheet["D35"]= rsChargeDetail[1]
        #cur.close()
        rsChargeDetails.close()
        #conn.close()
        return "Successful"        
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetChargeDetails " + str(e))
        return e
def format_seconds_to_hhmmss(seconds):
    hours = seconds / (60*60)
    seconds %= (60*60)
    minutes = seconds // 60
    seconds %= 60
    return "%02i:%02i:%02i" % (hours, minutes, seconds)

def mthdGetServiceSummary():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
        rsServiceDetail=[]
        xlParams.xlPosition = 8
        i= 0
        j=0
        xlParams.sheet = xlParams.workbook.worksheets[1]
        
         
        
        
        
        
        #style123= xw.easyxf('font: bold 1;height 300;align: horiz right')
        #style1234= xw.easyxf('font: bold 1;height 320;align: horiz right')
        rsServiceDetails=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_SERVICE_SUMMARY",[xlParams.lngInvoiceNo,rsServiceDetails])
        for row in rsServiceDetails:
            rsServiceDetail.append(row)
            ##print(rsServiceDetail)
        #print(len(rsServiceDetail))
        if(len(rsServiceDetail)>0):
            if(xlParams.xlFileName[0:3]=="CAR" or xlParams.xlFileName[0:2] =="PG" or xlParams.xlFileName[0:2]=="IC" ):
                xlParams.sheet["A7"] = "Phone Number or Circuit Id"
        while(i < len(rsServiceDetail)):
            total_net =0.00
            total_vat=0.00
            total_discnt =0.00
            tmp_header=""
            j =i
            tmp_header = rsServiceDetail[i][0]
            xlParams.sheet["A" + str(xlParams.xlPosition)] = rsServiceDetail[i][0]
            xlParams.sheet["A" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
            while(j < len(rsServiceDetail)):
                if(tmp_header == rsServiceDetail[i][0]):
                    if(rsServiceDetail[i][3] + rsServiceDetail[i][4] != 0):
                        #print(str(rsServiceDetail[i][10].strftime('%d/%m/%Y')) + "-" + str(rsServiceDetail[i][11].strftime('%d/%m/%Y')) )
                        #print(str(rsServiceDetail[i][11].strftime('%d/%m/%Y')))
                        if (rsServiceDetail[i][10] is not None  and rsServiceDetail[i][10] is not None ):
                            xlParams.sheet["C" + str(xlParams.xlPosition)] = (str(rsServiceDetail[i][10].strftime('%d/%m/%Y')) + " - " + str(rsServiceDetail[i][11].strftime('%d/%m/%Y')) )
                            xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                        xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False) 
                        xlParams.sheet["D" + str(xlParams.xlPosition)] = rsServiceDetail[i][3]/100
                        xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.sheet["E" + str(xlParams.xlPosition)]= rsServiceDetail[i][4]/100
                        xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.sheet["F" + str(xlParams.xlPosition)] = rsServiceDetail[i][5]/100
                        xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        xlParams.sheet["F" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["F" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        if(rsServiceDetail[i][6]/100<=86399):
                            xlParams.sheet["G" + str(xlParams.xlPosition)]= time.strftime('%H:%M:%S', time.gmtime(rsServiceDetail[i][6]/100))
                        else:
                            seconds=0
                            seconds=format_seconds_to_hhmmss(rsServiceDetail[i][6]/100)
                            xlParams.sheet["G" + str(xlParams.xlPosition)]= seconds
                        xlParams.sheet["G" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                        xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        total_net = total_net + rsServiceDetail[i][3]/100
                        total_vat = total_vat + rsServiceDetail[i][4]/100
                        total_discnt = total_discnt + rsServiceDetail[i][5]/100
                        if (rsServiceDetail[i][1] is not None or rsServiceDetail[i][1] ):
                            #print("The description is not null")
                            if(rsServiceDetail[i][2] == 63):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Local Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)]= "Local Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Local Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)]= "Local Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Local Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Local Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Local Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            elif(rsServiceDetail[i][2] == 55):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Here Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Here Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "Here Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Here Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Here Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Here Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Here Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            elif(rsServiceDetail[i][2] == 64):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Off-Peak Usage"                               
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="National Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            elif(rsServiceDetail[i][2] == 57):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="ROW Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] = "ROW Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            elif(rsServiceDetail[i][2] == 58 and rsServiceDetail[i][9] == 16 ):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="English Mobile/Inmarsat Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            elif(rsServiceDetail[i][2] == 58 and rsServiceDetail[i][9] == 12 ):
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] ="Irish Mobile Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                            else:
                                #print("the loop comes to else part")
                                if(rsServiceDetail[i][7] == "B"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Bank Holiday Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "W"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Weekend Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "D"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Daytime Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "E"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Evening Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "P"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                elif(rsServiceDetail[i][7] == "O"):
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =str(rsServiceDetail[i][1]) + " - Off-Peak Usage"
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                                else:
                                    #print("the loop comes to last else part")
                                    xlParams.sheet["B" + str(xlParams.xlPosition)] =rsServiceDetail[i][1]
                                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                        #print("The Loops comeout")                                         
                        xlParams.xlPosition = xlParams.xlPosition + 1
                    i = i + 1
                    j= i
                    
                else:
                    if(total_net !=0 or total_vat !=0 or total_discnt !=0):
                        #print("Entering Total Service Initial Loop")
                        xlParams.sheet["D" + str(xlParams.xlPosition)] =total_net
                        xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                        xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                        xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.sheet["E" + str(xlParams.xlPosition)] =total_vat
                        xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                        xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                        xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.sheet["F" + str(xlParams.xlPosition)] =total_discnt
                        xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                        xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                        xlParams.sheet["F" + str(xlParams.xlPosition)].number_format ='0.00'
                        xlParams.sheet["F" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.sheet["B" + str(xlParams.xlPosition)] ="Totals for Service:"
                        xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                        xlParams.xlPosition = xlParams.xlPosition + 2
                        
                    j= len(rsServiceDetail) + 1
        #print("The length is" + str(len(rsServiceDetail)))    
        if(len(rsServiceDetail) > 0):
            if(total_net !=0 or total_vat !=0 or total_discnt !=0):
                #print("Entering Total Service Final Initial Loop")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_net
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["E" + str(xlParams.xlPosition)] =total_vat
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["F" + str(xlParams.xlPosition)] =float(total_discnt)
                xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["F" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["F" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["B" + str(xlParams.xlPosition)] ="Totals for Service:"
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                
        
        return "Successful"
        #cur.close()
        rsServiceDetails.close()
        #conn.close()
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetServiceSummary " + str(e))
        return e

def mthdGetPaymentsAndAdj() :
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
         
        
        rsPaymentsAndAdj=[]
        adj_count=0
        total_adj = 0
        
        
        
        xlParams.xlPosition = xlParams.xlPosition + 2
        rsPaymentsAndAdjs=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_PAY_AND_ADJ",[xlParams.lngInvoiceNo,rsPaymentsAndAdjs])
        for row in rsPaymentsAndAdjs:
            rsPaymentsAndAdj.append(row)
            #print(rsPaymentsAndAdj)
        if(len(rsPaymentsAndAdj) > 0):
            xlParams.xlPosition = xlParams.xlPosition + 1
            xlParams.sheet["B" + str(xlParams.xlPosition)]  = "Adjustments"
            xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=15,bold=True)
            xlParams.sheet["B" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left", vertical="center") 
            xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["C" + str(xlParams.xlPosition)]  = ""
            xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left", vertical="center") 
            xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition)]  = ""
            xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left", vertical="center") 
            xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.xlPosition = xlParams.xlPosition + 1
            start_cell = xlParams.xlPosition
            #print("The Adjustments are available")
            while(adj_count < len(rsPaymentsAndAdj)):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =rsPaymentsAndAdj[adj_count][0]
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =rsPaymentsAndAdj[adj_count][1]
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                total_adj = float(total_adj) + float(rsPaymentsAndAdj[adj_count][1])
                adj_count =adj_count + 1
                xlParams.xlPosition = xlParams.xlPosition + 1
            #print(sum(rsPaymentsAndAdj[1]))
                
                
            if (xlParams.xlBillCurrency == 'IEP'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_adj
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'GBP'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total £"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_adj
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'USD'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total $"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_adj
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'EURO'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total €"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_adj
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            else:
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total €"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_adj
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
                
            #xlParams.sheet('B32' :'D34').fill = style1
        
        
        return "Successful"
        #cur.close()
        rsPaymentsAndAdjs.close()
        #conn.close()      
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetPaymentsAndAdj " + str(e))
        return e
        
def mthdGetCCs():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
         
        
        rsCCDetail=[]
        dis_count=0
        total_dis = 0
        
        
        
        rsCCDetails=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_CREDITS_AND_CHARGES",[xlParams.lngInvoiceNo,rsCCDetails])
        for ccdrow in rsCCDetails:
            rsCCDetail.append(ccdrow)
            #print(rsCCDetail)
        if(len(rsCCDetail) > 0):
            xlParams.sheet["B" + str(xlParams.xlPosition)] = "Other Credits and Charges"
            xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=15,bold=True)
            xlParams.sheet["B" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left")
            xlParams.xlPosition = xlParams.xlPosition + 1
            #print("The Discounts are available")
            while(dis_count < len(rsCCDetail)):
                xlParams.sheet["C" + str(xlParams.xlPosition)] =rsCCDetail[dis_count][0]
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =rsCCDetail[dis_count][1]
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                total_dis = float(total_dis) + float(rsCCDetail[dis_count][1])
                dis_count =dis_count + 1
                xlParams.xlPosition = xlParams.xlPosition + 1
            #print(sum(rsPaymentsAndAdj[1]))
                
                
            if (xlParams.xlBillCurrency == 'IEP'):
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right")
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_dis
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'GBP'):
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total £"
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_dis
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'USD'):
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total $"
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_dis
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'EURO'):
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total €"
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_dis
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
            else:
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total €"
                xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total"
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] =total_dis
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
                
        return "Successful"
        #cur.close()
        rsCCDetails.close()
        #conn.close()              
        
    
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetCCs " + str(e))
        return e
    
def mthdGetDiscounts():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
         
        
        
        
        
        rsDiscountsDetail=[]
        dis_count1=0
        DiscountTotal = 0
        
        rsDiscountsDetails=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_DISCOUNTS",[xlParams.lngInvoiceNo,rsDiscountsDetails])
        for disc in rsDiscountsDetails:
            rsDiscountsDetail.append(disc)
            #print(rsDiscountsDetail)
        xlParams.xlPosition = xlParams.xlPosition + 3
        #FixedPosition = xlParams.xlPosition
        
        if(len(rsDiscountsDetail) > 0):
            xlParams.sheet["B" + str(xlParams.xlPosition)] ="Discounts"
            xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=15,bold=True)
            xlParams.sheet["B" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left", vertical="center") 
            xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["C" + str(xlParams.xlPosition)] =""
            xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition)] =""
            xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["E" + str(xlParams.xlPosition)] =""
            xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.xlPosition = xlParams.xlPosition + 1
            while(dis_count1 < len(rsDiscountsDetail)):
                if(rsDiscountsDetail[dis_count1][1] != 0):
                    xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                    xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["C" + str(xlParams.xlPosition)] =(rsDiscountsDetail[dis_count1][0])
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["D" + str(xlParams.xlPosition)] =""
                    xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =(rsDiscountsDetail[dis_count1][1])/100
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                    xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    xlParams.xlPosition = xlParams.xlPosition + 1
                DiscountTotal = DiscountTotal + (rsDiscountsDetail[dis_count1][1])    
                dis_count1 = dis_count1 + 1
                
                
            
            xlParams.sheet["C" + str(xlParams.xlPosition)] =""
            xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition)] =""
            xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["E" + str(xlParams.xlPosition)] =""
            xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["B" + str(xlParams.xlPosition)] =""
            xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.xlPosition = xlParams.xlPosition + 1
            if (xlParams.xlBillCurrency == 'IEP'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =""
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] ="Total discount:"
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)] =DiscountTotal/100
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'GBP'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =""
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] ="Total discount £:"
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)] =DiscountTotal/100
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'USD'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =""
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] ="Total discount $:"
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)] =DiscountTotal/100
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            elif (xlParams.xlBillCurrency == 'EURO'):
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =""
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] ="Total discount €:"
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)] =DiscountTotal/100
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            else:
                xlParams.sheet["B" + str(xlParams.xlPosition)] =""
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)] =""
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)] ="Total discount €:"
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)] =DiscountTotal/100
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format ='0.00'
                xlParams.sheet["E" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 1
            
        return "Successful"
        #cur.close()
        rsDiscountsDetails.close()
        #conn.close()       
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetDiscounts " + str(e))
        return e
    
def mthdGetServiceSummaryTotals ():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
         
        
        
        rsServiceDetailTotal=[]
        xlParams.xlPosition = xlParams.xlPosition + 1
        t_cnt = 0
        t_cnt1 = 0
        t_xl_position = 30
        #dis_count1=0
        #DiscountTotal = 0
        
        rsServiceDetailTotals=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_SERVICE_SUMMARY_TOTALS",[xlParams.lngInvoiceNo,rsServiceDetailTotals])
        for rs_total in rsServiceDetailTotals:
            rsServiceDetailTotal=rs_total
        if(len(rsServiceDetailTotal) > 0):
            xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Discount Received"
            xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
            xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[4]
            xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
            xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format ='0.00'
            xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            xlParams.sheet["C" + str(xlParams.xlPosition + 1)] ="Total Before Discount and Tax"
            xlParams.sheet["C" + str(xlParams.xlPosition+ 1)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["C" + str(xlParams.xlPosition+ 1)].alignment  = Alignment(horizontal="right", vertical="center") 
            xlParams.sheet["C" + str(xlParams.xlPosition+ 1)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition + 1)] =rsServiceDetailTotal[2]
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].font = Font(name='Century Gothic',size=10,bold=True)
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].alignment  = Alignment(horizontal="right", vertical="center") 
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].number_format ='0.00'
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            if(rsServiceDetailTotal[3] != 0):
                if (xlParams.invoice_dt < datetime.datetime(2020, 9, 1)):
                    xlParams.sheet["C" + str(xlParams.xlPosition + 3)] ="Total VAT @ 23%"
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["D" + str(xlParams.xlPosition + 3)] =rsServiceDetailTotal[3]
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format ='0.00'
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    xlParams.xlPosition = xlParams.xlPosition + 3
                elif (xlParams.invoice_dt >= datetime.datetime(2020, 9, 1) and xlParams.invoice_dt < datetime.datetime(2021, 3, 1)):
                    xlParams.sheet["C" + str(xlParams.xlPosition + 3)] ="Total VAT @ 21%"
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["D" + str(xlParams.xlPosition + 3)] =rsServiceDetailTotal[3]
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format ='0.00'
                    xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    xlParams.xlPosition = xlParams.xlPosition + 3
                else :
                    taxtotals=xlParams.conn.cursor()
                    taxtotal=[]
                    #xlParams.logger.info("Loop 1 comes" )
                    xlParams.cur.callproc("ARBOR.Bed_Get_Tax_Details",[xlParams.lngInvoiceNo,0,xlParams.invoice_dt,1,taxtotals])
                    for tx_total in taxtotals:
                         taxtotal.append(tx_total)
                    if(len(taxtotal) > 0):
                        #xlParams.logger.info("Loop 2 comes" )
                        xlParams.xlPosition = xlParams.xlPosition + 2
                        while (t_cnt < taxtotal[0][0]):
                            #xlParams.logger.info(taxtotal[0][0])
                            #xlParams.logger.info(taxtotal[0][1])
                            xlParams.xlPosition = xlParams.xlPosition  + 1
                            #xlParams.logger.info("Loop 3 comes" )
                            xlParams.sheet["C" + str(xlParams.xlPosition)] ="Total VAT @ " + str(taxtotal[t_cnt][5]/10000) + str("%")
                            #xlParams.logger.info("Loop 4 comes" )
                            #xlParams.logger.info(taxtotal[t_cnt][2] )
                            #xlParams.logger.info(taxtotal[t_cnt][5] )
                            xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                            xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                            xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                            xlParams.sheet["D" + str(xlParams.xlPosition)] =taxtotal[t_cnt][2]/100
                            xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                            xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                            xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                            xlParams.sheet["D" + str(xlParams.xlPosition)].number_format ='0.00'
                            xlParams.sheet["D" + str(xlParams.xlPosition)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                            t_cnt = t_cnt + 1
                         
                    else:                      
                        xlParams.sheet["C" + str(xlParams.xlPosition + 3)] ="Total VAT @ 23%"
                        xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                        xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                        xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                        xlParams.sheet["D" + str(xlParams.xlPosition + 3)] =rsServiceDetailTotal[3]
                        xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                        xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                        xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                        xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format ='0.00'
                        xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        xlParams.xlPosition = xlParams.xlPosition + 3
                    
            else:
                xlParams.sheet["C" + str(xlParams.xlPosition + 3)] ="Total VAT"
                xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 3)] =rsServiceDetailTotal[3]
                xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                xlParams.xlPosition = xlParams.xlPosition + 3
                
            #xlParams.sheet["D" + str(xlParams.xlPosition + 3)] =rsServiceDetailTotal[3]
            #xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].font = Font(name='Century Gothic',size=10,bold=True)
            #xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].alignment  = Alignment(horizontal="right", vertical="center") 
            #xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].fill =PatternFill("solid", fgColor="00C0C0C0")
            #xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format ='0.00'
            #xlParams.sheet["D" + str(xlParams.xlPosition+ 3)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            xlParams.sheet["C" + str(xlParams.xlPosition + 1)] =""
            xlParams.sheet["C" + str(xlParams.xlPosition + 1)].fill =PatternFill("solid", fgColor="00C0C0C0")
            xlParams.sheet["D" + str(xlParams.xlPosition + 1)] =""
            xlParams.sheet["D" + str(xlParams.xlPosition+ 1)].fill =PatternFill("solid", fgColor="00C0C0C0")
            if (xlParams.xlBillCurrency == 'IEP'):
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Charge"
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[5]
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            elif (xlParams.xlBillCurrency == 'GBP'):
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Charge £"
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[5]
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            elif (xlParams.xlBillCurrency == 'USD'):
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Charge $"
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[5]
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            elif (xlParams.xlBillCurrency == 'EURO'):
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Charge €"
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[5]
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            else:
                xlParams.sheet["C" + str(xlParams.xlPosition + 2)] ="Total Charge €"
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["C" + str(xlParams.xlPosition+ 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition + 2)] =rsServiceDetailTotal[5]
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].alignment  = Alignment(horizontal="right", vertical="center") 
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format ='0.00'
                xlParams.sheet["D" + str(xlParams.xlPosition+ 2)].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        
        xlParams.sheet = xlParams.workbook.worksheets[0]         
        xlParams.sheet["D26"] =rsServiceDetailTotal[0]
        xlParams.sheet["D27"]  =rsServiceDetailTotal[1]
        xlParams.sheet["D29"] = rsServiceDetailTotal[2]
        xlParams.sheet["D30"]  = rsServiceDetailTotal[4]
        if(rsServiceDetailTotal[3] != 0):
            if (xlParams.invoice_dt < datetime.datetime(2020, 9, 1)):
                t_xl_position = t_xl_position + 1
                xlParams.sheet["C31"] = "Total VAT @ 23%"
                xlParams.sheet["D31"]  =rsServiceDetailTotal[3]
            elif (xlParams.invoice_dt >= datetime.datetime(2020, 9, 1) and xlParams.invoice_dt < datetime.datetime(2021, 3, 1)):
                t_xl_position = t_xl_position + 1
                xlParams.sheet["C31"] = "Total VAT @ 21%"
                xlParams.sheet["D31"]  =rsServiceDetailTotal[3]
            else :
                #xlParams.sheet["C31"] ="Total VAT @ 23%"
                if(len(taxtotal) > 0):
                    while (t_cnt1 < taxtotal[0][0]):
                        t_xl_position = t_xl_position + 1                    
                        xlParams.sheet["C" + str(t_xl_position)] ="Total VAT @ " + str(taxtotal[t_cnt1][5]/10000) + str("%")
                        xlParams.sheet["D" + str(t_xl_position)] =taxtotal[t_cnt1][2]/100
                        if (xlParams.xlBillCurrency =="GBP" or xlParams.xlBillCurrency =="USD"):
                            xlParams.sheet["E25"] ="Euro Vat Amt €"
                            xlParams.sheet["E25" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right")
                            xlParams.sheet["E25"].font = Font(name='Century Gothic',size=10,bold=False)
                            xlParams.sheet["E" + str(t_xl_position)] = round ((taxtotal[t_cnt1][7]/100),2)
                        t_cnt1 = t_cnt1 + 1
                else:
                    t_xl_position = t_xl_position + 1
                    xlParams.sheet["C31"] = "Total VAT @ 21%"
                    xlParams.sheet["D31"]  =rsServiceDetailTotal[3]



            if ((xlParams.xlBillCurrency =="GBP" or xlParams.xlBillCurrency =="USD") and  t_cnt1 <= 0):
                VatConversionRate=[]
                VatConversionAmount=0
                VatConversionRates=xlParams.conn.cursor()
                xlParams.cur.callproc("BOD_QUERY_TYPES.GET_VAT_CURR_CONV_RATE",[xlParams.xlBillCurrency,'EUR',str(xlParams.invoice_dt.strftime('%d%m%Y')),VatConversionRates])
                for v_rate in VatConversionRates:
                    VatConversionRate=v_rate
            
                if(len(VatConversionRate)>=0):
                    VatConversionAmount = round((VatConversionRate[0] * rsServiceDetailTotal[3]),2)
                    xlParams.sheet["E25"] ="Euro Vat Amt €"
                    xlParams.sheet["E25" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right")
                    xlParams.sheet["E25"].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E31"] = VatConversionAmount
                    xlParams.sheet["E31" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right")
                else:
                    return "No Conversion rate"
        else:
            t_xl_position = t_xl_position + 1
            xlParams.sheet["C31"]  ="Total VAT"
            xlParams.sheet["D31"]  =rsServiceDetailTotal[3]
        #xlParams.sheet["D31"]  =rsServiceDetailTotal[3]
        #xlParams.sheet["D34"]  =rsServiceDetailTotal[5] +  rsServiceDetailTotal[7]
        

        #xlParams.sheet["D32"]  =rsServiceDetailTotal[7]
        xlParams.sheet["C" + str(t_xl_position + 1)]  ="Credits incl VAT"
        xlParams.sheet["D" + str(t_xl_position + 1)]  =rsServiceDetailTotal[7]
        xlParams.sheet["D" + str(t_xl_position + 3)]  =rsServiceDetailTotal[5] +  rsServiceDetailTotal[7]
        xlParams.sheet["C" + str(t_xl_position + 3)] = "Invoice Total"
        xlParams.sheet["C" + str(t_xl_position + 3)].font = Font(name='Century Gothic',size=11,bold=True)
        #xlParams.sheet["D39"]  =rsServiceDetailTotal[6]
        if (rsServiceDetailTotal[0] != 0 or rsServiceDetailTotal[1] != 0 or rsServiceDetailTotal[7] != 0):
            xlParams.logger.info("Ths status is zero " )
            return "Successful"
        else:
            xlParams.logger.info("Ths status is Successful " )     
            return "Zero Invoice"
        #cur.close()
        rsDiscountsDetails.close()
        #conn.close()
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdGetServiceSummaryTotals " + str(e))
        return e
        
        
def mthdOutputDetails():
    try:
        #conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        #cur = conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
        xlParams.xlPosition =9
        
        
        
        
        rsServiceDetailTotal=[]
        
        table_name = "BOD_" + str(xlParams.lngInvoiceNo)
        xlParams.sheet = xlParams.workbook.worksheets[2]
        sheet3 = xlParams.workbook.worksheets[3]
        s_count = xlParams.cur.var(cx_Oracle.NUMBER)
        #rsServiceDetailTotals=conn.cursor()
        xlParams.cur.callproc("BOD_GET_CALL_DETAIL_ANALYSIS",[xlParams.lngInvoiceNo,s_count])  
        #print(s_count.getvalue(pos=0))
        if(s_count.getvalue(pos=0)<=0):
            #print("No Usage_FOUND")
            xlParams.sheet["B10"] = 'No usage found for this period'
            xlParams.sheet["B10"].font = Font(name='Century Gothic',size=11,bold=False)
            xlParams.sheet["B26"] = 'No usage found for this period'
            xlParams.sheet["B26"].font = Font(name='Century Gothic',size=11,bold=False)
            xlParams.sheet["B41"] = 'No usage found for this period'
            xlParams.sheet["B41"].font = Font(name='Century Gothic',size=11,bold=False)
            sheet3["B7"] = 'No usage found for this period'
            sheet3["B7"].font = Font(name='Century Gothic',size=11,bold=False)
        else :
            #print("Usage_FOUND")
            xlParams.cur.execute("""select "No Of Calls", POINT_TARGET, DESTINATION, TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || 
            TO_CHAR(MOD(Seconds,60),'FM00') TIME,AMOUNT from (SELECT count(SERVICE_INSTANCE) "No Of Calls", POINT_TARGET, DESTINATION, 
            sum(seconds) seconds ,sum(AMOUNT/100) AS AMOUNT
            FROM  {} group by POINT_TARGET,DESTINATION
            ORDER BY Count(SERVICE_INSTANCE) DESC) WHERE ROWNUM <= 10""".format(table_name))
            
            counter =0
            list_by_freq = []
            for result_by_freq in xlParams.cur.fetchall():
                list_by_freq.append(result_by_freq)
                if counter < 10:
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_freq[counter][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_freq[counter][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)] = list_by_freq[counter][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)] = list_by_freq[counter][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["G" + str(xlParams.xlPosition)] = list_by_freq[counter][4]
                    xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter = counter + 1
                    
            ### query four Top Ten calls by Revenue
            xlParams.xlPosition = 25
            list_by_rev = []
            xlParams.cur.execute("""select CALL_DATE, POINT_TARGET, DESTINATION, TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || 
            TO_CHAR(MOD(Seconds,60),'FM00') TIME,AMOUNT from (SELECT  CALL_DATE, POINT_TARGET, DESTINATION, 
            seconds ,AMOUNT/100 AS AMOUNT FROM  {}
            ORDER BY AMOUNT/100 DESC) WHERE ROWNUM <= 10""".format(table_name))
            counter =0
            for result_by_rev in xlParams.cur.fetchall():
                list_by_rev.append(result_by_rev)
                if counter < 10:
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_rev[counter][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_rev[counter][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_rev[counter][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_rev[counter][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["G" + str(xlParams.xlPosition)] =list_by_rev[counter][4]
                    xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter = counter + 1
                    
            ### query four Top Ten calls by Duration
            xlParams.cur.execute("""select CALL_DATE, POINT_TARGET, DESTINATION, TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || 
            TO_CHAR(MOD(Seconds,60),'FM00') TIME,AMOUNT from (SELECT  CALL_DATE, POINT_TARGET, DESTINATION, 
            seconds ,AMOUNT/100 AS AMOUNT
            FROM  {} ORDER BY seconds DESC) WHERE ROWNUM <= 10""".format(table_name))
            counter =0
            list_by_dur = []
            
            xlParams.xlPosition = 40
            for result_by_dur in xlParams.cur.fetchall():
                list_by_dur.append(result_by_dur)
                if counter < 10:
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_dur[counter][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_dur[counter][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_dur[counter][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_dur[counter][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["G" + str(xlParams.xlPosition)] =list_by_dur[counter][4]
                    xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter = counter + 1
                    
            ### query for call Analysis By Jurisdiction
            xlParams.sheet = xlParams.workbook.worksheets[3]
            xlParams.xlPosition = 5
            
            
            
            
            xlParams.sheet.column_dimensions[get_column_letter(1)].width = len("Service Instance")
            xlParams.sheet.column_dimensions[get_column_letter(2)].width = len("Analysis By Jurisdiction")
            xlParams.sheet.column_dimensions[get_column_letter(3)].width = len("Jurisdiction         ")
            xlParams.sheet.column_dimensions[get_column_letter(4)].width = len("No. of Calls")
            xlParams.sheet.column_dimensions[get_column_letter(5)].width = len("Duration  ")
            xlParams.sheet.column_dimensions[get_column_letter(6)].width = len("Cost €  ")
            xlParams.sheet.column_dimensions[get_column_letter(7)].width = len("Jurisdiction")
            xlParams.sheet.column_dimensions[get_column_letter(8)].width = len("Rate Period")
            
            xlParams.cur.execute("""select "Call Jurisdiction","Number",TO_CHAR(TRUNC(Seconds/3600)) || ':' ||
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || 
            TO_CHAR(MOD(Seconds,60),'FM00') Duration ,Cost from (SELECT 
            (regexp_replace(CALL_JURISDICTION,'[0-9]')  ||' ' ||  CALL_RATE_PERIOD )AS "Call Jurisdiction",
            Count(SERVICE_INSTANCE) "Number", Sum(Seconds) AS Seconds, Sum(AMOUNT/100) AS Cost FROM {}
            GROUP BY (regexp_replace(CALL_JURISDICTION,'[0-9]')  ||' ' ||  CALL_RATE_PERIOD )
            ORDER BY (regexp_replace(CALL_JURISDICTION,'[0-9]')  ||' ' ||  CALL_RATE_PERIOD ))""".format(table_name))
            counter_jur =0
            list_by_juris = []
            xlParams.xlPosition=8
            #rows_jur=0
            rows_jur = xlParams.cur.fetchall()
            xlParams.logger.debug("Rows returned by call Analysis By Jurisdiction = " + str(len(rows_jur)))
            if (len(rows_jur) >0):
                xlParams.sheet["A6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B6"]= 'Analysis By Jurisdiction'
                xlParams.sheet["B6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B6"].font = Font(name='Century Gothic',size=11,bold=True)                
                xlParams.sheet["C6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["F6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["G6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["H6"].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["I6"].fill =PatternFill("solid", fgColor="00C0C0C0")                               
                xlParams.sheet["C8"]= 'Jurisdiction'
                xlParams.sheet["C8"].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D8"]= 'No. of Calls'
                xlParams.sheet["D8"].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["E8"]= 'Duration'
                xlParams.sheet["E8"].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["F8"]= 'Cost €'
                xlParams.sheet["F8"].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
                for result_by_juris in rows_jur:
                    list_by_juris.append(result_by_juris)
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_juris[counter_jur][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_juris[counter_jur][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_juris[counter_jur][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_juris[counter_jur][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center") 
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter_jur = counter_jur + 1
            ### query for call Analysis By Destination
            xlParams.xlPosition= xlParams.xlPosition + 3
            
            xlParams.cur.execute("""select DESTINATION,"Number", TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || 
            TO_CHAR(MOD(Seconds,60),'FM00') TIME,Cost from (SELECT DESTINATION, Count(SERVICE_INSTANCE) AS "Number",
            Sum(Seconds) AS Seconds, Sum(AMOUNT/100) AS Cost FROM {} GROUP BY DESTINATION
            ORDER BY DESTINATION)""".format(table_name))
            counter_dest = 0
            list_by_dest = []
            #rows_dest =0
            rows_dest = xlParams.cur.fetchall()
            xlParams.logger.debug("Rows returned by call Analysis By Destination = " + str(len(rows_dest)))
            if (len(rows_dest)>0):
                xlParams.sheet["A" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B" + str(xlParams.xlPosition)]='Analysis By Destination'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["F" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["G" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["H" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["I" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.xlPosition=xlParams.xlPosition + 2
                xlParams.sheet["C" + str(xlParams.xlPosition)]='Destination'
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)]='No. of Calls'
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)]='Duration'
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["F" + str(xlParams.xlPosition)]='Cost €'
                xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.xlPosition=xlParams.xlPosition + 1
                for result_by_dest in rows_dest:
                    list_by_dest.append(result_by_dest)
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_dest[counter_dest][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left", vertical="center")
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_dest[counter_dest][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_dest[counter_dest][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_dest[counter_dest][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter_dest = counter_dest + 1
                
            ### query for call Analysis By Hour
            xlParams.xlPosition= xlParams.xlPosition + 3
            
            xlParams.cur.execute("""select "Hour","Number",TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || TO_CHAR(MOD(Seconds,60),'FM00') TIME,
            Cost from (SELECT substr(CALL_DATE,InStr(CALL_DATE,':')-2,2) "Hour", Count(SERVICE_INSTANCE) "Number", 
            Sum(Seconds) AS Seconds, Sum(AMOUNT/100) AS Cost FROM {} GROUP BY substr(CALL_DATE,InStr(CALL_DATE,':')-2,2)
            ORDER BY substr(CALL_DATE,InStr(CALL_DATE,':')-2,2))""".format(table_name))
            counter_hour = 0
            list_by_hour = []
            
            #rows_ansis =0
            rows_ansis = xlParams.cur.fetchall()
            xlParams.logger.debug("Rows returned by query for call Analysis By Hour = " + str(len(rows_ansis)))
            if (len(rows_ansis) > 0):
                xlParams.sheet["A" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B" + str(xlParams.xlPosition)]= 'Analysis By Hour'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["F" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["G" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["H" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["I" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.xlPosition = xlParams.xlPosition + 2
                xlParams.sheet["C" + str(xlParams.xlPosition)]='Hour'
                #xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="left")
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)]='No. of Calls'
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)]='Duration'
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["F" + str(xlParams.xlPosition)]='Cost €'
                xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
                for result_by_hour in rows_ansis:
                    list_by_hour.append(result_by_hour)
                    xlParams.sheet["C" + str(xlParams.xlPosition)] = list_by_hour[counter_hour][0]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["C" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["D" + str(xlParams.xlPosition)] = list_by_hour[counter_hour][1]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_hour[counter_hour][2]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_hour[counter_hour][3]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter_hour = counter_hour + 1
                    
            ### query for Calls Over €5
            xlParams.xlPosition= xlParams.xlPosition + 3
            
            xlParams.cur.execute("""SELECT SERVICE_INSTANCE, CALL_DATE, POINT_ORIGIN, POINT_TARGET, DESTINATION, TO_CHAR(TRUNC(Seconds/3600))
            || ':' || TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || TO_CHAR(MOD(Seconds,60),'FM00') TIME,
            AMOUNT/100 AS Expr1, CALL_JURISDICTION,  CALL_RATE_PERIOD
            FROM {} WHERE (((AMOUNT/100)>4.99)) ORDER BY AMOUNT/100 DESC""".format(table_name))
            counter_OVER5 = 0
            list_by_OVER5 = []
            #xlParams.xlPosition = xlParams.xlPosition + 1
            #rows_over5 =0
            rows_over5 = xlParams.cur.fetchall()
            xlParams.logger.debug("Rows returned by query for call Analysis over €5  = " + str(len(rows_over5)))
            if(len(rows_over5) > 0):
                xlParams.sheet["A" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B" + str(xlParams.xlPosition)] ='Calls Over €5'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["F" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["G" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["H" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["I" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.xlPosition = xlParams.xlPosition + 2
                xlParams.sheet["A" + str(xlParams.xlPosition)] ='Service Instance'
                xlParams.sheet["A" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)] ='Call Date & Time'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)] ='Origin'
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] ='Target'
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)] ='Destination'
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["F" + str(xlParams.xlPosition)] ='Duration'
                xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["G" + str(xlParams.xlPosition)] ='Cost €'
                xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["H" + str(xlParams.xlPosition)] ='Jurisdiction'
                xlParams.sheet["H" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["I" + str(xlParams.xlPosition)] ='Rate Period'
                xlParams.sheet["I" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
                for result_by_over5 in rows_over5:
                    list_by_OVER5.append(result_by_over5)
                    xlParams.sheet["A" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][0]
                    xlParams.sheet["A" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["B" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][1]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["B" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["C" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][2]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][3]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][4]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][5]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["G" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][6]
                    xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["G" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["H" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][7]
                    xlParams.sheet["H" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["I" + str(xlParams.xlPosition)] =list_by_OVER5[counter_OVER5][8]
                    xlParams.sheet["I" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter_OVER5 = counter_OVER5 + 1
                
            ### query for Calls Over 30 Minutes
            xlParams.xlPosition= xlParams.xlPosition + 3
            
            
            xlParams.cur.execute("""SELECT SERVICE_INSTANCE, CALL_DATE, POINT_ORIGIN, POINT_TARGET, DESTINATION, TO_CHAR(TRUNC(Seconds/3600)) || ':' || 
            TO_CHAR(TRUNC(MOD(Seconds,3600)/60),'FM00') || ':' || TO_CHAR(MOD(Seconds,60),'FM00') TIME, AMOUNT/100 AS Expr1,
            CALL_JURISDICTION,  CALL_RATE_PERIOD FROM {} WHERE (Seconds)>1799 ORDER BY Seconds DESC""".format(table_name))
            counter_OVER30Min = 0
            list_by_OVER30Min = []
            #xlParams.xlPosition = xlParams.xlPosition + 1
            rows_over30 =0
            rows_over30 = xlParams.cur.fetchall()
            xlParams.logger.debug("Rows returned by query for call Analysis over 30 min = " + str(len(rows_over30)))
            if (len(rows_over30) > 0) :
                xlParams.sheet["A" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["B" + str(xlParams.xlPosition)] ='Calls Over 30 Minutes'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=11,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["C" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["D" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["E" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["F" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["G" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["H" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.sheet["I" + str(xlParams.xlPosition)].fill =PatternFill("solid", fgColor="00C0C0C0")
                xlParams.xlPosition = xlParams.xlPosition + 2
                xlParams.sheet["A" + str(xlParams.xlPosition)] ='Service Instance'
                xlParams.sheet["A" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["B" + str(xlParams.xlPosition)] ='Call Date & Time'
                xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["C" + str(xlParams.xlPosition)] ='Origin'
                xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["D" + str(xlParams.xlPosition)] ='Target'
                xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["E" + str(xlParams.xlPosition)] ='Destination'
                xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["F" + str(xlParams.xlPosition)] ='Duration'
                xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["G" + str(xlParams.xlPosition)] ='Cost €'
                xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["H" + str(xlParams.xlPosition)] ='Jurisdiction'
                xlParams.sheet["H" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.sheet["I" + str(xlParams.xlPosition)] ='Rate Period'
                xlParams.sheet["I" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=True)
                xlParams.xlPosition = xlParams.xlPosition + 1
                
                
                for result_by_OVER30Min in rows_over30:
                    list_by_OVER30Min.append(result_by_OVER30Min)
                    xlParams.sheet["A" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][0]
                    xlParams.sheet["A" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["B" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][1]
                    xlParams.sheet["B" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["B" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["C" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][2]
                    xlParams.sheet["C" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["D" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][3]
                    xlParams.sheet["D" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["E" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][4]
                    xlParams.sheet["E" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][5]
                    xlParams.sheet["F" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["F" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["G" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][6]
                    xlParams.sheet["G" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["G" + str(xlParams.xlPosition)].alignment  = Alignment(horizontal="right", vertical="center")
                    xlParams.sheet["H" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][7]
                    xlParams.sheet["H" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.sheet["I" + str(xlParams.xlPosition)] =list_by_OVER30Min[counter_OVER30Min][8]
                    xlParams.sheet["I" + str(xlParams.xlPosition)].font = Font(name='Century Gothic',size=10,bold=False)
                    xlParams.xlPosition = xlParams.xlPosition + 1
                    counter_OVER30Min = counter_OVER30Min + 1
                 
        
        #file_location = xlParams.Root_Directory + Template_directory + 'laxman.png' 
        drop_table_status = drop_table()
        
         
        xlParams.workbook.save(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + '.xlsx') 
        return "Successful"    
        xlParams.cur.close()
        #rsDiscountsDetails.close()
        xlParams.conn.close()
        
        
        
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdOutputDetails " + str(e))
        return e

def drop_table():
    try:
        xlParams.cur.callproc("BOD_DROP_CALL_DETAIL_TABLE",[xlParams.lngInvoiceNo])
        return "Successful"

    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  writeinto_csv " + str(e))
        return e

def writeinto_csv():
    try:
        xlParams.conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        xlParams.cur = xlParams.conn.cursor()
        #xlParams.lngInvoiceNo = 26186859
         
        
        
        rsServiceDetailTotal=[]
        rsServiceDetail_Header = [["CLI","CALL_DATE","POINT_ORIGIN","POINT_TARGET","Destination","Second","AMOUNT","JURISDICTION","RATE_PERIOD"]]
        #dis_count1=0
        #DiscountTotal = 0
        
        rsServiceDetailTotals=xlParams.conn.cursor()
        xlParams.cur.callproc("bod_create_types.BOD_GET_CALL_DETAILS",[xlParams.lngInvoiceNo,rsServiceDetailTotals])
        for row in rsServiceDetailTotals:
            rsServiceDetailTotal.append(row)
        if(len(rsServiceDetailTotal) > 0):
            csv.register_dialect('myDialect',delimiter=',',quoting=csv.QUOTE_ALL)
            with open(xlParams.Prod_directory + xlParams.xlFileName + ".csv", 'w', newline='') as file:
                writer = csv.writer(file, dialect='myDialect')
                writer.writerows(rsServiceDetail_Header)
                writer.writerows(rsServiceDetailTotal)
        else :
            csv.register_dialect('myDialect',delimiter=',',quoting=csv.QUOTE_ALL)
            with open(xlParams.Prod_directory + xlParams.xlFileName + ".csv", 'w', newline='') as file:
                writer = csv.writer(file, dialect='myDialect')
                writer.writerows(rsServiceDetail_Header)
                #writer.writerows(rsServiceDetailTotal)
            
        return "Successful"
        cur.close()
        rsServiceDetailTotals.close()
        conn.close()
    
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  writeinto_csv " + str(e))
        return e
def Generte_Random_Password(plength):
    # get random password pf length 8 with letters, digits, and symbols
    characters = string.ascii_letters + string.digits #+ string.punctuation
    password = ''.join(random.choice(characters) for i in range(plength))
    print("The password is "+ password)
    return password
def zip_files(cnt):
    try:
        if cnt==1:
            mail_password = str(Generte_Random_Password(10))
            pass_word=  '-p' + mail_password
                    #cmd = 'C:/Program Files/7-Zip/7za a -tzip -p"passw0rd" -mem=ZipCrypto C:\\Users\\606545301\\BOD\\Production\\11-E270E3-1-26489925-01062021.ZIP C:\\Users\\606545301\\BOD\\PasswordProtect\\11-E270E3-1-26489925-01062021.ZIP'
            rc = subprocess.call([xlParams.zip_directory, 'a', '-mem=ZipCrypto', pass_word, '-y', xlParams.Password_directory + xlParams.xlFileName + ".zip"] + 
                                [xlParams.Prod_directory + xlParams.xlFileName + ".xlsx"])

            print("The password is "+ pass_word)
            return "Successful" ,mail_password
        else:
            mail_password = str(Generte_Random_Password(10))
            pass_word=  '-p' + mail_password
                    #cmd = 'C:/Program Files/7-Zip/7za a -tzip -p"passw0rd" -mem=ZipCrypto C:\\Users\\606545301\\BOD\\Production\\11-E270E3-1-26489925-01062021.ZIP C:\\Users\\606545301\\BOD\\PasswordProtect\\11-E270E3-1-26489925-01062021.ZIP'
            rc = subprocess.call([xlParams.zip_directory, 'a', '-mem=ZipCrypto', pass_word, '-y', xlParams.Password_directory + xlParams.xlFileName + ".zip"] + 
                                [xlParams.Prod_directory + xlParams.xlFileName + ".xlsx" ,xlParams.Prod_directory + xlParams.xlFileName + ".csv"])

            print("The password is "+ pass_word)
            return "Successful",mail_password
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  zip_files " + str(e))
        return e ,mail_password
def send_mail(zip_type,cust_email,zipname,PASSWORD):
    #log_setup()
    #current_dir = os. getcwd()
    file_csv = ""
    if (zip_type==1):
        #os.chdir(xlParams.Root_Directory +  xlParams.Password_directory)
        #file_name=  zipname + ".zip"
        #file_name=os.path.basename(xlParams.Root_Directory +  xlParams.Password_directory + zipname + ".zip")
        file_name=xlParams.Root_Directory +  xlParams.Password_directory + zipname + ".zip"
        #print(file_name)
    else:
        if (path.exists(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx") and path.exists(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv") ):
            file_csv ="True"
            file_name=xlParams.Root_Directory +  xlParams.Prod_directory + zipname + ".xlsx"
            file_name1=xlParams.Root_Directory +  xlParams.Prod_directory + zipname + ".csv"
        else:
            file_name=xlParams.Root_Directory +  xlParams.Prod_directory + zipname + ".xlsx"
        #file_name=  zipname + ".zip"
        #print(file_name)

    #zf = open(file_name,'rb')
    msg = MIMEMultipart()
    msg['From'] = xlParams.MailFromAddress
    msg['To'] = cust_email
    msg['Subject'] = xlParams.MailSubject + " " + str(xlParams.lngInvoiceNo)
    if (zip_type==1):
        message = xlParams.encryptMailBody
        part2 = MIMEText(message, 'html')
        msg.attach(part2)
    else:
        message = xlParams.MailBody
        part2 = MIMEText(message, 'html')
        msg.attach(part2)

    msg_pwd = MIMEMultipart()
    msg_pwd['From'] = xlParams.MailFromAddress
    msg_pwd['To'] = cust_email
    msg_pwd['Subject'] = xlParams.PasswordMailSubject + " " + str(xlParams.lngInvoiceNo)
    message_pwd = xlParams.PasswordMailBody
    #print("The password is " +str(PASSWORD))
    message_pwd1=message_pwd.replace("PASSWORD",PASSWORD)
    #print(message_pwd)
    part2_pwd = MIMEText(message_pwd1, 'html')
    #part2_pwd.replace("PASSWORD",PASSWORD)
    msg_pwd.attach(part2_pwd)

    with open(file_name,'rb') as file:
    # Attach the file with filename to the email
        msg.attach(MIMEApplication(file.read(), Name=os.path.basename(file_name)))
    if file_csv =="True":
        with open(file_name1,'rb') as file:
            msg.attach(MIMEApplication(file.read(), Name=os.path.basename(file_name1)))

    smtp_server = xlParams.MailServer
    port = 25  # For starttls
    sender_email = xlParams.MailFromAddress
    #print("The mail server is" +str(smtp_server) )
    #password = input("Type your password and press enter: ")

    # Create a secure SSL context
    #context = ssl.create_default_context()

    # Try to log in to server and send email
    try:
        server = smtplib.SMTP(smtp_server,port)
        server.ehlo() # Can be omitted
        server.starttls() # Secure the connection
        server.ehlo() # Can be omitted
        #server.login(sender_email, password)
        # TODO: 
        if (zip_type==1):
            server.sendmail(sender_email,cust_email,str(msg))
            server.sendmail(sender_email,cust_email,str(msg_pwd))
        else:
            server.sendmail(sender_email,cust_email,str(msg))

       # os.chdir(current_dir)
        return "Successful"


    except Exception as e:
        # Print any error messages to stdout
        #os.chdir(current_dir)
        xlParams.logger.warning("There is an Exception sending mail "+ str(e))
        return e
        #print(e)
    finally:
        server.quit()
#send_mail(1,'lakshminarayanareddy.pittu@bt.com','99073012183045-2-26547708-01072021') 

def mthdSaveExcelSheet(bool_csv,cust_email):
    try:
        list_files=[]
        status =""
        mailing_password=""
        if(bool_csv=="True"):
            if (path.exists(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx") and path.exists(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv") ):                           
                list_files = [xlParams.Prod_directory + xlParams.xlFileName + ".xlsx" ,xlParams.Prod_directory + xlParams.xlFileName + ".csv"]                
                xlParams.logger.info("Yes both  CSV and Excel files" + xlParams.xlFileName + ".xlsx  are Exists" )
                with zipfile.ZipFile (xlParams.Prod_directory + xlParams.xlFileName + ".zip",'w') as zipF:
                    for file in list_files:
                        zipF.write(file ,compress_type=zipfile.ZIP_DEFLATED)
                if(xlParams.encrypt=="True" ):                           
                    zip_status,mailing_password = zip_files(2)
                    xlParams.logger.info("The status of Encryption method is "+ str(zip_status))                    
                    
                    if(zip_status=="Successful"):
                        if(len(cust_email)>=5):
                            send_mail_status=send_mail(1,cust_email,xlParams.xlFileName,mailing_password) 
                            xlParams.logger.info("The status of Sending mail is "+ str(send_mail_status))
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv")
                            status = send_mail_status
                        else:
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv")
                            status ="Successful"
                    else:
                        status ="Failure"
                else:
                    
                    if(len(cust_email)>=5):
                            send_mail_status=send_mail(2,cust_email,xlParams.xlFileName,'') 
                            xlParams.logger.info("The status of Sending mail is "+ str(send_mail_status))
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv")
                            status = send_mail_status
                    else:
                        os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                        os.remove(xlParams.Root_Directory +xlParams.Prod_directory + xlParams.xlFileName + ".csv")
                        status ="Successful"


            else :
                xlParams.logger.warning("No the files" + xlParams.Root_Directory +  xlParams.xlFileName + " .xlsx is not available" )
                status ="Failure"
        else :
            if (path.exists( xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")):                             
                list_files = [xlParams.Prod_directory + xlParams.xlFileName + ".xlsx"]                
                xlParams.logger.info("Yes the Excel file" + xlParams.Root_Directory + xlParams.xlFileName + ".xlsx is Available" )
                with zipfile.ZipFile ( xlParams.Prod_directory + xlParams.xlFileName + ".zip",'w') as zipF:
                    for file in list_files:
                        zipF.write(file ,compress_type=zipfile.ZIP_DEFLATED)
                if(xlParams.encrypt=="True" ):
                    zip_status,mailing_password = zip_files(1)
                    xlParams.logger.info("The status of Encryption method is "+ str(zip_status))
                    
                    if(zip_status=="Successful"):
                        if(len(cust_email)>=5):
                            send_mail_status=send_mail(1,cust_email,xlParams.xlFileName,mailing_password)
                            xlParams.logger.info("The status of Sending mail is "+ str(send_mail_status)) 
                            os.remove(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            status = send_mail_status
                        else:
                            os.remove(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            status ="Successful"
                    else:
                        status ="Failure"
                        
                       
                else:
                    
                    if(len(cust_email)>=5):
                            send_mail_status=send_mail(2,cust_email,xlParams.xlFileName,'') 
                            xlParams.logger.info("The status of Sending mail is "+ str(send_mail_status))
                            os.remove(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                            status = send_mail_status
                    else:
                        os.remove(xlParams.Root_Directory + xlParams.Prod_directory + xlParams.xlFileName + ".xlsx")
                        status ="Successful"

                
            else :
                (xlParams.logger.error("No the Excel file " + xlParams.Root_Directory + xlParams.xlFileName + " .xlsx is not available" ))
                status ="Failure"
                
        return status
        
        
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  mthdSaveExcelSheet " + str(e))
        return e
        
def close_logging():
    try:
        xlParams.file_handler.close()
        xlParams.logger.removeHandler(xlParams.file_handler)
        del xlParams.file_handler
        handlers = xlParams.logger.handlers[:]
        for handler in handlers:
            handler.close()
            xlParams.logger.removeHandler(handler)
        return "Successful"
    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  close_logging " + str(e))
        return e
    
        
def create_pdf(cust_email,user_name,boolUsesEncrypt):
    try:
        xlParams.conn = cx_Oracle.connect(xlParams.Db_username,xlParams.Db_Password,xlParams.Db_Database)
        xlParams.cur = xlParams.conn.cursor()
        #rsCustomerDetails=conn.cursor()
        
        xlParams.cur.callproc("BOD_BATCH_JOBS_PKG.BOD_INSERT_PDF_INVOICE",[user_name,xlParams.lngInvoiceNo,cust_email,boolUsesEncrypt])
        #print(rsCustomerDetails)
        xlParams.cur.close()
        #xlParams.rsServiceDetailTotals.close()
        xlParams.conn.close()       

        return "Successful"

    except Exception as e: 
        xlParams.logger.warning("We have an exception in method  create_pdf " + str(e))
        return e
    
        
        
    

#mthdGetCustInvoice_status = mthdGetCustInvoice ('99112211102551-2','26323844',1,0,1,'lakshminarayanareddy.pittu@bt.com','laxman')
#print(mthdGetCustInvoice_status)

